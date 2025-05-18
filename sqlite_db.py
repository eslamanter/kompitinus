import os
import sqlite3
import sys
import bcrypt
import config
import pandas as pd
from openpyxl import load_workbook, Workbook
from PyQt5.QtCore import QDateTime
from utils import exists, show_warning_msg
from constants import (DB_USERS_ID_BASE, DB_USERS_TABLE, DB_USER_ID,
                       DB_FIRST_NAME, DB_LAST_NAME, DB_EMAIL, DB_PIN, DB_REGISTERED_AT, DB_ACTIVE, DB_TASKS_ID_BASE,
                       DB_TASKS_TABLE, DB_TASK_ID, DB_SENDER_ID, DB_RECEIVER_ID, DB_CREATED_AT, DB_MODIFIED_AT,
                       DB_TITLE, DB_BODY, DB_REFERENCE, DB_DUE_AT, DB_STARRED, DB_DONE, DB_EXPECTED_AT, DB_REPLY,
                       DB_ARCHIVED, CFG_PATH, DB_SEEN_AT, UI_INBOX, UI_OUTBOX, UI_EXPIRED_BOX, UI_STARRED_BOX,
                       UI_TASK_ID, UI_MODIFIED_AT, UI_SENDER, UI_RECEIVER, UI_TASK, UI_DUE_AT, DB_REPORT_VIEW, UI_DELAY,
                       UI_DONE, MSG_EXCEL_OPENED, MSG_DB_UNREACHABLE, ERR_DB)


def export_report_view():
    """Exports DB report view into a new Excel worksheet of a new or an existing workbook in the local directory."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to the database
            conn = sqlite3.connect(config.config[CFG_PATH])

            # Query the view
            query = f"SELECT * FROM {DB_REPORT_VIEW}"
            df = pd.read_sql_query(query, conn)

            # Define Excel file path
            excel_path = f"{DB_REPORT_VIEW}.xlsx"

            # Generate a sheet name using current date and time
            sheet_name = QDateTime.currentDateTime().toString("yyyy-MM-dd_HH-mm-ss")

            # Check if file exists
            if exists(excel_path):
                # Load existing workbook
                wb = load_workbook(excel_path)
            else:
                # Create new workbook and remove default sheet
                wb = Workbook()
                default_sheet = wb.active
                wb.remove(default_sheet)
                ws = wb.create_sheet(title=sheet_name)  # Create new sheet as the first sheet

            # If loading an existing file, create a new sheet
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)

            # Write dataframe to the new sheet
            for r_idx, row in enumerate(df.values, start=2):  # Start from row 2 to leave space for headers
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Add headers manually
            for c_idx, header in enumerate(df.columns, start=1):
                ws.cell(row=1, column=c_idx, value=header)

            # Add filters to the first row
            ws.auto_filter.ref = ws.dimensions

            # List of columns to adjust
            columns = ["A", "B", "C", "D", "E", "F", "G", "H"]

            # Adjust column width based on content
            for col_letter in columns:
                col_idx = ws[col_letter]
                max_length = max((len(str(cell.value)) for cell in col_idx if cell.value), default=0)
                ws.column_dimensions[col_letter].width = max_length + 2  # Adding padding

            # Set the last created worksheet as the active sheet
            wb.active = len(wb.sheetnames) - 1

            # Save the updated Excel file
            wb.save(excel_path)

            # Close connection
            conn.close()
            os.startfile(excel_path)
            return True
        except PermissionError:
            show_warning_msg(MSG_EXCEL_OPENED)
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def get_task_details(task_id):
    """Retrieves all task details given the task ID."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Query to fetch the task by ID
            cursor.execute(f"SELECT * FROM {DB_TASKS_TABLE} WHERE {DB_TASK_ID} = ?", (task_id,))

            # Fetch the result
            task_data = cursor.fetchone()

            # Close the connection
            conn.close()

            return task_data
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def get_tasks_by_user(user_id, box_type, filter_type=None):
    """Retrieves all tasks of a user given their ID, box type, and optionally box filter."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        sql_conditions = ""
        params = []

        # Boxes
        if box_type == UI_INBOX:
            sql_conditions = f"WHERE t.{DB_RECEIVER_ID} = ? "
            sql_conditions += f"AND t.{DB_ARCHIVED} = 0 "
            params.append(user_id)

        elif box_type == UI_OUTBOX:
            sql_conditions = f"WHERE t.{DB_SENDER_ID} = ? "
            params.append(user_id)

        # Filters
            # Starred filter
        if filter_type == UI_STARRED_BOX:
            sql_conditions += f"AND t.{DB_STARRED} = 1 "

            # Expired filter
        elif filter_type == UI_EXPIRED_BOX:
            current_time = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            sql_conditions += f"AND t.{DB_DONE} = 0 "
            if box_type == UI_OUTBOX:
                sql_conditions += f"AND t.{DB_ARCHIVED} = 0 "
            sql_conditions += f"AND t.{DB_DUE_AT} < ? "
            params.append(current_time)

            # Other user filter
        elif config.my_id != user_id:
            sql_conditions += f"AND t.{DB_STARRED} = 1 "
            sql_conditions += f"AND t.{DB_DONE} = 0 "

        # Last line of SQL conditions
        if filter_type == UI_EXPIRED_BOX:
            sql_conditions += f"ORDER BY t.{DB_DUE_AT}"
        else:
            sql_conditions += f"ORDER BY t.{DB_MODIFIED_AT} DESC"

        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            cursor.execute(f"""
                SELECT 
                    t.{DB_TASK_ID}, 
                    t.{DB_MODIFIED_AT}, 
                    t.{DB_STARRED}, 
                    t.{DB_ARCHIVED}, 
                    t.{DB_TITLE}, 
                    t.{DB_DUE_AT}, 
                    t.{DB_DONE}, 
                    sender.{DB_FIRST_NAME} AS sender_first_name, 
                    sender.{DB_LAST_NAME} AS sender_last_name,
                    receiver.{DB_FIRST_NAME} AS receiver_first_name, 
                    receiver.{DB_LAST_NAME} AS receiver_last_name
                FROM {DB_TASKS_TABLE} AS t
                JOIN {DB_USERS_TABLE} AS sender ON t.{DB_SENDER_ID} = sender.{DB_USER_ID}
                JOIN {DB_USERS_TABLE} AS receiver ON t.{DB_RECEIVER_ID} = receiver.{DB_USER_ID}
                {sql_conditions}
            """, tuple(params))

            return cursor.fetchall()
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def add_task(sender_id, receiver_id, title, body, reference, due_at, expected_at, starred, archived, reply="", done=0):
    """Adds new task to DB given all task details and returns its ID."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Insert query
            cursor.execute(f"""
                INSERT INTO {DB_TASKS_TABLE} (
                    {DB_SENDER_ID}, {DB_RECEIVER_ID}, {DB_TITLE}, {DB_BODY}, 
                    {DB_REFERENCE}, {DB_DUE_AT}, {DB_EXPECTED_AT}, {DB_STARRED},
                    {DB_ARCHIVED}, {DB_REPLY}, {DB_DONE}
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (sender_id, receiver_id, title, body, reference, due_at, expected_at, starred, archived, reply, done))

            # Commit changes to save the insertion
            conn.commit()
            conn.close()
            return cursor.lastrowid
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def update_task(title, body, reference, due_at, expected_at, starred, archived, reply, done, task_id):
    """Updates an existing task given its ID and details."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Update query
            cursor.execute(f"""
                UPDATE {DB_TASKS_TABLE}
                SET {DB_TITLE} = ?, {DB_BODY} = ?, {DB_REFERENCE} = ?, {DB_DUE_AT} = ?, {DB_EXPECTED_AT} = ?,
                    {DB_STARRED} = ?, {DB_ARCHIVED} = ?, {DB_REPLY} = ?, {DB_DONE} = ?,
                    {DB_MODIFIED_AT} = datetime('now', 'localtime')
                WHERE {DB_TASK_ID} = ?
            """, (title, body, reference, due_at, expected_at,
                  starred, archived, reply, done,
                  task_id))

            # Commit changes to save the update
            conn.commit()
            conn.close()
            return True
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def get_all_users():
    """Retrieves firstname, lastname, and email of all active users."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Fetch all users, ordered by last name then first name
            cursor.execute(f"""
            SELECT {DB_USER_ID}, {DB_FIRST_NAME}, {DB_LAST_NAME}, {DB_EMAIL}
            FROM {DB_USERS_TABLE}
            WHERE {DB_ACTIVE} = 1
            ORDER BY {DB_LAST_NAME} ASC, {DB_FIRST_NAME} ASC
            """)

            return cursor.fetchall()  # Get all rows
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    sys.exit() # Exists if DB connection is not possible when app initializes UI


def email_exists(email):
    """Checks if given email exists."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Execute the query
            cursor.execute(F"SELECT 1 FROM {DB_USERS_TABLE} WHERE {DB_EMAIL} = ?", (email,))
            result = cursor.fetchone()

            # Close the connection
            conn.close()

            return result
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def get_my_seen_at():
    """Retrieves local user last seen at."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to the database
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            cursor.execute(f"""
            SELECT {DB_SEEN_AT} FROM {DB_USERS_TABLE}
            WHERE {DB_USER_ID} = ?
            """, (config.my_id,))
            result = cursor.fetchone()
            return result[0] if result else None  # Returns the timestamp or None if the user isn't found
        except Exception as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def get_user_email(user_id):
    """Retrieves email for a given user ID."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to the database
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Query to fetch user details
            cursor.execute(f"""
                SELECT {DB_EMAIL}
                FROM {DB_USERS_TABLE}
                WHERE {DB_USER_ID} = ?
            """, (user_id,))

            # Fetch the result
            user_info = cursor.fetchone()

            # Close connection
            conn.close()

            # Return the user info if found
            return user_info[0] if user_info else None

        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def get_user_full_name(user_id):
    """Retrieves firstname and lastname for a given user ID."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to the database
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Query to fetch user details
            cursor.execute(f"""
                SELECT {DB_FIRST_NAME}, {DB_LAST_NAME}
                FROM {DB_USERS_TABLE}
                WHERE {DB_USER_ID} = ?
            """, (user_id,))

            # Fetch the result
            user_info = cursor.fetchone()

            # Close connection
            conn.close()

            # Return the user info if found
            return user_info if user_info else None

        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


def add_new_user(first_name, last_name, email, pin):
    """Adds new user given firstname, lastname, email, and pin and assigns ID to config.my_id."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Hash user pin
            salt = bcrypt.gensalt()
            hashed_pin = bcrypt.hashpw(pin.encode(), salt)

            # User data to insert
            user_data = (first_name, last_name, email, hashed_pin,)

            # Insert query
            cursor.execute(f"""
                INSERT INTO {DB_USERS_TABLE} ({DB_FIRST_NAME}, {DB_LAST_NAME}, {DB_EMAIL}, {DB_PIN})
                VALUES (?, ?, ?, ?)
            """, user_data)

            # Get user ID
            config.my_id = cursor.lastrowid

            # Commit and close
            conn.commit()
            conn.close()
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")


def update_my_seen_at():
    """Updates local user last seen at."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Update query
            cursor.execute(f"""
                    UPDATE {DB_USERS_TABLE}
                    SET {DB_SEEN_AT} = (datetime('now', 'localtime'))
                    WHERE {DB_USER_ID} = ?
                """, (config.my_id,))

            # Commit and close
            conn.commit()
            conn.close()
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")


def update_user_data(first_name, last_name, email, pin):
    """Updates user data given new: firstname, lastname, email, and pin."""
    if not exists(config.config[CFG_PATH]):
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Hash user pin
            salt = bcrypt.gensalt()
            hashed_pin = bcrypt.hashpw(pin.encode(), salt)

            # Update query
            cursor.execute(f"""
                UPDATE {DB_USERS_TABLE}
                SET {DB_FIRST_NAME} = ?, {DB_LAST_NAME} = ?, {DB_EMAIL} = ?, {DB_PIN} = ?
                WHERE {DB_USER_ID} = ?
            """, (first_name, last_name, email, hashed_pin, config.my_id))

            # Commit and close
            conn.commit()
            conn.close()
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")


def check_login(email, pin):
    """Verifies login by checking email-based user ID and hashed pin."""
    if not exists(config.config[CFG_PATH]):  # Ensure database exists
        show_warning_msg(MSG_DB_UNREACHABLE)
    else:
        try:
            # Connect to main DB
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Retrieve user ID and hashed PIN from registered email
            cursor.execute(f"SELECT {DB_USER_ID}, {DB_PIN} FROM {DB_USERS_TABLE} WHERE {DB_EMAIL} = ?",
                           (email,))
            result = cursor.fetchone()

            if result:
                user_id, hashed_pin = result  # Extract user ID and hashed PIN
                raw = pin.encode('utf-8')  # Convert input PIN to bytes

                # Securely check hashed PIN
                if bcrypt.checkpw(raw, hashed_pin):
                    config.my_id = user_id
                    return True  # Successful login

                return False  # Incorrect PIN
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")

    return None  # Missing/Inaccessible DB or invalid email


def create_db():
    """Creates new DB only if it does not exist."""
    if not exists(config.config[CFG_PATH]):
        try:
            # Create and connect to SQLite database
            conn = sqlite3.connect(config.config[CFG_PATH])
            cursor = conn.cursor()

            # Create users table
            cursor.execute(f"""
            CREATE TABLE {DB_USERS_TABLE} (
                {DB_USER_ID}    INTEGER UNIQUE,
                {DB_FIRST_NAME}     TEXT NOT NULL,
                {DB_LAST_NAME}      TEXT NOT NULL,
                {DB_EMAIL}          TEXT NOT NULL UNIQUE,
                {DB_PIN}            BLOB NOT NULL,
                {DB_SEEN_AT}        TEXT DEFAULT (datetime('now', 'localtime')),
                {DB_REGISTERED_AT}  TEXT DEFAULT (datetime('now', 'localtime')),
                {DB_ACTIVE}         INTEGER DEFAULT 1,
                PRIMARY KEY({DB_USER_ID} AUTOINCREMENT)
            );
            """)

            # Create tasks table
            cursor.execute(f"""
            CREATE TABLE {DB_TASKS_TABLE} (
                {DB_TASK_ID}		INTEGER UNIQUE,
                {DB_SENDER_ID}		INTEGER NOT NULL,
                {DB_RECEIVER_ID}	INTEGER NOT NULL,
                {DB_CREATED_AT}	    TEXT DEFAULT (datetime('now', 'localtime')),
                {DB_MODIFIED_AT}	TEXT DEFAULT (datetime('now', 'localtime')),
                {DB_TITLE}			TEXT,
                {DB_BODY}			TEXT,
                {DB_REFERENCE}		TEXT,
                {DB_DUE_AT}		    TEXT,
                {DB_STARRED}		INTEGER,
                {DB_DONE}		    INTEGER,
                {DB_EXPECTED_AT}	TEXT,
                {DB_REPLY}			TEXT,
                {DB_ARCHIVED}		INTEGER,
                PRIMARY KEY({DB_TASK_ID} AUTOINCREMENT),
                FOREIGN KEY({DB_SENDER_ID}) REFERENCES {DB_USERS_TABLE}({DB_USER_ID}),
                FOREIGN KEY({DB_RECEIVER_ID}) REFERENCES {DB_USERS_TABLE}({DB_USER_ID})
            );
            """)

            # Create report view
            cursor.execute(f"""
                CREATE VIEW {DB_REPORT_VIEW} AS
                    SELECT
                        {DB_TASKS_TABLE}.{DB_TASK_ID} AS {UI_TASK_ID.replace(" ", "_")},
                        {DB_TASKS_TABLE}.{DB_MODIFIED_AT} AS {UI_MODIFIED_AT.replace(" ", "_")},
                        CONCAT(sender.{DB_FIRST_NAME}, ' ', sender.{DB_LAST_NAME}) AS {UI_SENDER.replace(" ", "_")},
                        CONCAT(receiver.{DB_FIRST_NAME}, ' ', receiver.{DB_LAST_NAME}) AS {UI_RECEIVER.replace(" ", "_")},
                        {DB_TASKS_TABLE}.{DB_TITLE} AS {UI_TASK.replace(" ", "_")},
                        {DB_TASKS_TABLE}.{DB_DUE_AT} AS {UI_DUE_AT.replace(" ", "_")},
                        {DB_TASKS_TABLE}.{DB_DONE} AS {UI_DONE.replace(" ", "_")},
    
                        CASE
                            WHEN {DB_TASKS_TABLE}.{DB_DONE} = 0 THEN
                                CAST((julianday('now') - julianday({DB_TASKS_TABLE}.{DB_DUE_AT})) AS INTEGER)
                            ELSE NULL
                        END AS {UI_DELAY}
    
                    FROM
                        {DB_TASKS_TABLE}
                    JOIN {DB_USERS_TABLE} AS sender ON {DB_TASKS_TABLE}.{DB_SENDER_ID} = sender.{DB_USER_ID}
                    JOIN {DB_USERS_TABLE} AS receiver ON {DB_TASKS_TABLE}.{DB_RECEIVER_ID} = receiver.{DB_USER_ID}
                    WHERE {DB_TASKS_TABLE}.{DB_STARRED} = 1 AND {DB_TASKS_TABLE}.{DB_ARCHIVED} = 0;
                """)

            # Update sqlite_sequence
            dummy_text = "dummy_text"
            dummy_int = DB_USERS_ID_BASE

            cursor.execute(
                f"INSERT INTO {DB_USERS_TABLE} ({DB_FIRST_NAME}, {DB_LAST_NAME}, {DB_EMAIL}, {DB_PIN}) VALUES ('{dummy_text}', '{dummy_text}', '{dummy_text}', '{dummy_text}')")
            cursor.execute(f"DELETE FROM {DB_USERS_TABLE} WHERE {DB_EMAIL} = '{dummy_text}'")

            cursor.execute(
                f"INSERT INTO {DB_TASKS_TABLE} ({DB_SENDER_ID}, {DB_RECEIVER_ID}, {DB_TITLE}) VALUES ({dummy_int}, {dummy_int}, '{dummy_text}')")
            cursor.execute(f"DELETE FROM {DB_TASKS_TABLE} WHERE {DB_TITLE} = '{dummy_text}'")

            cursor.execute(f"UPDATE sqlite_sequence SET seq = {DB_USERS_ID_BASE} WHERE name = '{DB_USERS_TABLE}'")
            cursor.execute(f"UPDATE sqlite_sequence SET seq = {DB_TASKS_ID_BASE} WHERE name = '{DB_TASKS_TABLE}'")

            # Commit changes and close connection
            conn.commit()
            conn.close()

            return True
        except sqlite3.Error as e:
            show_warning_msg(f"{ERR_DB}: {str(e)}")
    return False


if __name__ == "__main__":
    create_db()